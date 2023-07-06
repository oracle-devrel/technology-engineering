# Copyright (c) 2023, Oracle and/or its affiliates.
# Licensed under the Universal Permissive License v1.0 as shown at https://oss.oracle.com/licenses/upl.
# Purpose: Read WLS MPI template Excel, extract the variables and generate JSON variable files. 
# @author: Vasudeva Manikandan

import json
import io
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

# load the workbook
wb = load_workbook('wl_mpi_template_v2.xlsx')
file1 = open('input_v3.json', 'w')
file1.write('{\n')
# Get a sheet by name
for sheets in wb.sheetnames:
  sheet=wb[sheets]
  i=0
  for row in sheet.values:
     if (i>=1 and row[6] is not None):
       file1.write('"' + str(row[0]) + '":"' + str(row[6]) + '"' + ',\n')
     i = i + 1

file1.write('}')
file1.close()
